import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── DataFrames ────────────────────────────────────────────────────────────────
#synthetic data for testing, not actual data
# Each DataFrame is populated from email extraction logic but not included here for privacy reasons.

zelle_df = pd.DataFrame(columns=[
    "Date", "Amount", "Sender Name", "Memo"
])

cashapp_df = pd.DataFrame(columns=[
    "Date", "Amount", "Sender Name", "Cashtag", "Transaction ID", "Note"
])

paypal_df = pd.DataFrame(columns=[
    "Date", "Gross Amount", "Fee", "Net Amount",
    "Sender Name", "Sender Email", "Transaction ID", "Description", "Status"
])

# ─── Sample rows (remove once real data is loaded) ────────────────────────────
zelle_df = pd.concat([zelle_df, pd.DataFrame([
    {"Date": "2024-01-07", "Amount": 50.00,  "Sender Name": "John Smith",    "Memo": "January offering"},
    {"Date": "2024-01-14", "Amount": 100.00, "Sender Name": "Mary Johnson",  "Memo": "Tithe"},
    {"Date": "2024-01-21", "Amount": 75.00,  "Sender Name": "David Brown",   "Memo": "Building fund"},
])], ignore_index=True)

cashapp_df = pd.concat([cashapp_df, pd.DataFrame([
    {"Date": "2024-01-07", "Amount": 25.00,  "Sender Name": "Alice Green",  "Cashtag": "$alicegreen",  "Transaction ID": "CA001", "Note": "Offering"},
    {"Date": "2024-01-14", "Amount": 60.00,  "Sender Name": "Robert White", "Cashtag": "$rwhite",      "Transaction ID": "CA002", "Note": "Tithe"},
    {"Date": "2024-01-21", "Amount": 40.00,  "Sender Name": "Susan Clark",  "Cashtag": "$sclark",      "Transaction ID": "CA003", "Note": "Missions"},
])], ignore_index=True)

paypal_df = pd.concat([paypal_df, pd.DataFrame([
    {"Date": "2024-01-07", "Gross Amount": 100.00, "Fee": 3.20, "Net Amount": 96.80,
     "Sender Name": "Michael Hall",  "Sender Email": "mhall@email.com",  "Transaction ID": "PP001", "Description": "Donation", "Status": "Completed"},
    {"Date": "2024-01-14", "Gross Amount": 200.00, "Fee": 6.10, "Net Amount": 193.90,
     "Sender Name": "Patricia Young", "Sender Email": "pyoung@email.com", "Transaction ID": "PP002", "Description": "Tithe",    "Status": "Completed"},
    {"Date": "2024-01-21", "Gross Amount": 50.00,  "Fee": 1.75, "Net Amount": 48.25,
     "Sender Name": "James King",    "Sender Email": "jking@email.com",   "Transaction ID": "PP003", "Description": "Offering", "Status": "Completed"},
])], ignore_index=True)


# ─── Excel Export ──────────────────────────────────────────────────────────────

HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT     = Font(name="Arial", size=10)
TOTAL_FONT    = Font(name="Arial", bold=True, size=10)

THIN_BORDER   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color="888888"))

TAB_CONFIG = {
    "Zelle":   {"df": zelle_df,   "header_color": "4A235A", "tab_color": "4A235A", "amount_cols": ["Amount"]},
    "CashApp": {"df": cashapp_df, "header_color": "00C244", "tab_color": "00C244", "amount_cols": ["Amount"]},
    "PayPal":  {"df": paypal_df,  "header_color": "003087", "tab_color": "003087", "amount_cols": ["Gross Amount", "Fee", "Net Amount"]},
}

ALT_ROW_FILL = PatternFill("solid", start_color="F5F5F5")


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame,
                header_color: str, tab_color: str, amount_cols: list):
    ws = wb.create_sheet(title=name)
    ws.sheet_properties.tabColor = tab_color

    header_fill = PatternFill("solid", start_color=header_color)
    cols = list(df.columns)

    # Header row
    for c_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30

    # Data rows
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = ALT_ROW_FILL if r_idx % 2 == 0 else PatternFill()
        for c_idx, (col, val) in enumerate(zip(cols, row), start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if col in amount_cols:
                cell.number_format = '$#,##0.00;($#,##0.00);"-"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col == "Date":
                cell.number_format = "YYYY-MM-DD"
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Totals row
    total_row = len(df) + 2

    for c_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=total_row, column=c_idx)
        cell.font = TOTAL_FONT
        cell.border = BOTTOM_BORDER
        if col in amount_cols and not df.empty:
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            cell.number_format = '$#,##0.00;($#,##0.00);"-"'
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill = PatternFill("solid", start_color="E8E8E8")
        elif c_idx == 1:
            cell.value = "TOTAL"
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = PatternFill("solid", start_color="E8E8E8")
        else:
            cell.fill = PatternFill("solid", start_color="E8E8E8")

    # Auto-fit column widths
    for c_idx, col in enumerate(cols, start=1):
        max_len = len(str(col))
        for r_idx in range(2, total_row + 1):
            val = ws.cell(row=r_idx, column=c_idx).value
            if val and not str(val).startswith("="):
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"


def build_combined_df() -> pd.DataFrame:
    """Combine Zelle, CashApp, and PayPal into one standardized DataFrame."""
    zelle = pd.DataFrame({
        "Date":           pd.to_datetime(zelle_df["Date"], errors="coerce"),
        "Platform":       "Zelle",
        "Sender":         zelle_df["Sender Name"],
        "Amount":         pd.to_numeric(zelle_df["Amount"], errors="coerce").fillna(0.0),
        "Note":           zelle_df["Memo"],
        "Transaction_ID": pd.NA,
    })
    cashapp = pd.DataFrame({
        "Date":           pd.to_datetime(cashapp_df["Date"], errors="coerce"),
        "Platform":       "Cash App",
        "Sender":         cashapp_df["Sender Name"],
        "Amount":         pd.to_numeric(cashapp_df["Amount"], errors="coerce").fillna(0.0),
        "Note":           cashapp_df["Note"],
        "Transaction_ID": cashapp_df["Transaction ID"],
    })
    paypal = pd.DataFrame({
        "Date":           pd.to_datetime(paypal_df["Date"], errors="coerce"),
        "Platform":       "PayPal",
        "Sender":         paypal_df["Sender Name"],
        "Amount":         pd.to_numeric(paypal_df["Net Amount"], errors="coerce").fillna(0.0),
        "Note":           paypal_df["Description"],
        "Transaction_ID": paypal_df["Transaction ID"],
    })
    combined = pd.concat([zelle, cashapp, paypal], ignore_index=True)
    combined = combined.dropna(subset=["Date"]).sort_values("Date").reset_index(drop=True)
    combined["Running_Total"] = combined["Amount"].cumsum()
    return combined


def build_period_summary(combined_df: pd.DataFrame, period: str) -> pd.DataFrame:
    """Aggregate combined transactions to weekly ('W'), monthly ('M'), or yearly ('Y') totals."""
    period_start = combined_df["Date"].dt.to_period(period).dt.start_time
    summary = (
        combined_df.assign(Period_Start=period_start)
        .groupby("Period_Start", as_index=False)
        .agg(
            Transactions=("Amount", "count"),
            Period_Total=("Amount", "sum"),
        )
        .sort_values("Period_Start")
        .reset_index(drop=True)
    )
    summary["Running_Total"] = summary["Period_Total"].cumsum()
    return summary


SUMMARY_CONFIG = {
    "Combined_Report": {
        "header_color": "2C3E50", "tab_color": "2C3E50",
        "amount_cols": ["Amount", "Running_Total"],
    },
    "Weekly_Summary": {
        "header_color": "1A5276", "tab_color": "1A5276",
        "amount_cols": ["Period_Total", "Running_Total"],
    },
    "Monthly_Summary": {
        "header_color": "145A32", "tab_color": "145A32",
        "amount_cols": ["Period_Total", "Running_Total"],
    },
    "Yearly_Summary": {
        "header_color": "6E2F9E", "tab_color": "6E2F9E",
        "amount_cols": ["Period_Total", "Running_Total"],
    },
}


def export_to_excel(output_path: str):
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # Individual platform sheets
    for name, cfg in TAB_CONFIG.items():
        write_sheet(
            wb, name,
            df=cfg["df"],
            header_color=cfg["header_color"],
            tab_color=cfg["tab_color"],
            amount_cols=cfg["amount_cols"],
        )

    # Combined and period summary sheets
    combined = build_combined_df()
    summary_sheets = {
        "Combined_Report": combined,
        "Weekly_Summary":  build_period_summary(combined, "W"),
        "Monthly_Summary": build_period_summary(combined, "M"),
        "Yearly_Summary":  build_period_summary(combined, "Y"),
    }
    for name, df in summary_sheets.items():
        cfg = SUMMARY_CONFIG[name]
        write_sheet(
            wb, name,
            df=df,
            header_color=cfg["header_color"],
            tab_color=cfg["tab_color"],
            amount_cols=cfg["amount_cols"],
        )

    wb.save(output_path)
    print(f"Exported: {output_path}")


if __name__ == "__main__":
    export_to_excel("offerings_export.xlsx")
